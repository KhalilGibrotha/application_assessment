#!/bin/python3 python3
# -*- coding: utf-8 -*-

import openpyxl

# create a list of application instances
def create_app():
    # prompt the user for input
    name = input("Enter the name of the application: ")
    criticality = input("Enter the criticality of the application ({}): ".format(", ".join(CRITICALITIES)))
    availability = input("Enter the availability of the application ({}): ".format(", ".join(AVAILABILITIES)))
    number_of_users = input("Enter the number of users of the application: ")
    business_function = input("Enter the business function of the application ({}): ".format(", ".join(BUSINESS_FUNCTIONS)))
    environments = input("Enter the environments of the application ({}): ".format(", ".join(ENVIRONMENTS)))
    servers = input("Enter the number of servers of the application: ")
    memory = input("Enter the memory of the application: ")
    storage = input("Enter the storage of the application: ")
    database = input("Enter the database of the application ({}): ".format(", ".join(DATABASES)))
    operating_system = input("Enter the operating system of the application ({}): ".format(", ".join(OPERATING_SYSTEMS)))
    programming_language = input("Enter the programming language of the application ({}): ".format(", ".join(PROGRAMMING_LANGUAGES)))
    framework = input("Enter the framework of the application ({}): ".format(", ".join(FRAMEWORKS)))
    version_control = input("Enter the version control of the application ({}): ".format(", ".join(VERSION_CONTROL)))
    deployment = input("Enter the deployment of the application ({}): ".format(", ".join(DEPLOYMENT)))
    monitoring = input("Enter the monitoring of the application ({}): ".format(", ".join(MONITORING)))
    logging = input("Enter the logging of the application ({}): ".format(", ".join(LOGGING)))
    security_zones = input("Enter the security zones of the application ({}): ".format(", ".join(SECURITY_ZONES)))
    backup = input("Enter the backup of the application ({}): ".format(", ".join(BACKUP)))
    disaster_recovery = input("Enter the disaster recovery of the application ({}): ".format(", ".join(DISASTER_RECOVERY)))
    cost = input("Enter the cost of the application: ")
    revenue = input("Enter the revenue of the application: ")
    profit = input("Enter the profit of the application: ")
    notes = input("Enter any notes about the application: ")


class Application:
    def __init__(self, name, criticality, availability, number_of_users, business_owners, business_function, environments, servers, memory, storage, database, operating_system, programming_language, framework, version_control, deployment, monitoring, logging, security_zones, backup, disaster_recovery, cost, revenue, profit, notes):
        self.name = name
        self.criticality = criticality
        self.availability = availability
        self.number_of_users = number_of_users
        self.business_owners = business_owners
        self.business_function = business_function
        self.environments = environments
        self.servers = servers
        self.memory = memory
        self.storage = storage
        self.database = database
        self.operating_system = operating_system
        self.programming_language = programming_language
        self.framework = framework
        self.version_control = version_control
        self.deployment = deployment
        self.monitoring = monitoring
        self.logging = logging
        self.security_zones = security_zones
        self.backup = backup
        self.disaster_recovery = disaster_recovery
        self.cost = cost
        self.revenue = revenue
        self.profit = profit
        self.notes = notes

    
    # write the application data to the excel spreadsheet
    def write_to_excel(self, ws, row_num):
        ws.cell(row=row_num, column=1).value = self.name
        ws.cell(row=row_num, column=2).value = self.criticality
        ws.cell(row=row_num, column=3).value = self.availability
        ws.cell(row=row_num, column=4).value = self.number_of_users
        ws.cell(row=row_num, column=5).value = self.business_function
        ws.cell(row=row_num, column=6).value = self.environments
        ws.cell(row=row_num, column=7).value = self.servers
        ws.cell(row=row_num, column=8).value = self.memory
        ws.cell(row=row_num, column=9).value = self.storage
        ws.cell(row=row_num, column=10).value = self.database
        ws.cell(row=row_num, column=11).value = self.operating_system
        ws.cell(row=row_num, column=12).value = self.programming_language
        ws.cell(row=row_num, column=13).value = self.framework
        ws.cell(row=row_num, column=14).value = self.version_control
        ws.cell(row=row_num, column=15).value = self.deployment
        ws.cell(row=row_num, column=16).value = self.monitoring
        ws.cell(row=row_num, column=17).value = self.logging
        ws.cell(row=row_num, column=18).value = self.security_zones
        ws.cell(row=row_num, column=19).value = self.backup
        ws.cell(row=row_num, column=20).value = self.disaster_recovery
        ws.cell(row=row_num, column=21).value = self.cost
        ws.cell(row=row_num, column=22).value = self.revenue
        ws.cell(row=row_num, column=23).value = self.profit
        ws.cell(row=row_num, column=24).value = self.notes

if __name__ == "__main__":
    # open the excel spreadsheet
    wb = openpyxl.load_workbook('applications.xlsx')
    ws = wb.active

    # create a list of application instances
    applications = []

    # read the excel spreadsheet and create the application instances
    for row in ws.iter_rows(min_row=2, max_col=24, values_only=True):
        applications.append(Application(*row))

    # write the application instances to the excel spreadsheet
    for i, application in enumerate(applications):
        application.write_to_excel(ws, i+2)

    # save the excel spreadsheet
    wb.save('applications.xlsx')



