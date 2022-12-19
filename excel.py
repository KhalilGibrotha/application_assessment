#!/bin/python3
# -*- coding: utf-8 -*-
#
import openpyxl

def generate():
    # create a new workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # create a list of application instances
    applications = [
        Application("App 1", "High", "99.9%", 1000, "Finance", ["Prod", "Dev"], 2, 16, 500, "MySQL", "Linux",
                    "Python", "Django", "Git", "Ansible", "New Relic", "Splunk", ["DMZ", "Intranet"], "AWS S3",
                    "AWS RDS", 100000, 200000, 100000, "Some notes"),
        Application("App 2", "Low", "99.5%", 100, "HR", ["Prod"], 1, 8, 250, "PostgreSQL", "Windows", "Java",
                    "Spring", "SVN", "Jenkins", "New Relic", "Splunk", ["Intranet"], "AWS S3", "AWS RDS", 50000,
                    150000, 100000, "Other notes"),
        Application("App 3", "Medium", "99.0%", 500, "Marketing", ["Prod", "QA"], 4, 16, 500, "Oracle", "Linux",
                    "Java", "Hibernate", "Git", "Ansible", "New Relic", "Splunk", ["DMZ", "Intranet"], "AWS S3",
                    "AWS RDS", 75000, 150000, 75000, "")]

    # write each application's information to a row in the worksheet
    row_num = 1
    for app in applications:
        app.write_to_excel(ws, row_num)
        row_num += 1

    # save the workbook
    wb.save("applications.xlsx")


def update_app():
    # open the existing workbook and worksheet
    wb = openpyxl.load_workbook("applications.xlsx")
    ws = wb.active

    # create a list to store the application instances
    applications = []

    # iterate over the rows in the worksheet
    for row in ws.iter_rows(min_row=2, values_only=True):
        # unpack the row values into variables
        name, criticality, availability, number_of_users, business_function, environments, servers, memory, storage, database, operating_system, programming_language, framework, version_control, deployment, monitoring, logging, security_zones, backup, disaster_recovery, cost, revenue, profit, notes = row

        # prompt the user for any missing data
        if not name:
            name = input("Enter the name of the application: ")
        if not criticality:
            criticality = input("Enter the criticality of the application ({}): ".format(", ".join(CRITICALITIES)))
        if not availability:
            availability = input("Enter the availability of the application ({}): ".format(", ".join(AVAILABILITIES)))
        if not number_of_users:
            number_of_users = input("Enter the number of users: ")
        if not business_function:
            business_function = input("Enter the business function ({}): ".format(", ".join(BUSINESS_FUNCTIONS)))
        if not environments:
            environments = input("Enter the environments ({}): ".format(", ".join(ENVIRONMENTS)))
        if not servers:
            servers = input("Enter the number of servers: ")
        if not memory:
            memory = input("Enter the memory: ")
        if not storage:
            storage = input("Enter the storage: ")
        if not database:
            database = input("Enter the database ({}): ".format(", ".join(DATABASES)))
        if not operating_system:
            operating_system = input("Enter the operating system ({}): ".format(", ".join(OPERATING_SYSTEMS)))
        if not programming_language:
            programming_language = input("Enter the programming language ({}): ".format(", ".join(PROGRAMMING_LANGUAGES)))
        if not framework:
            framework = input("Enter the framework ({}): ".format(", ".join(FRAMEWORKS)))
        if not version_control:
            version_control = input("Enter the version control ({}): ".format(", ".join(VERSION_CONTROL)))
        if not deployment:
            deployment = input("Enter the deployment ({}): ".format(", ".join(DEPLOYMENT)))
        if not monitoring:
            monitoring = input("Enter the monitoring ({}): ".format(", ".join(MONITORING)))
        if not logging:
            logging = input("Enter the logging ({}): ".format(", ".join(LOGGING)))
        if not security_zones:
            security_zones = input("Enter the security zones ({}): ".format(", ".join(SECURITY_ZONES)))
        if not backup:
            backup = input("Enter the backup ({}): ".format(", ".join(BACKUP)))
        if not disaster_recovery:
            disaster_recovery = input("Enter the disaster recovery ({}): ".format(", ".join(DISASTER_RECOVERY)))
        if not cost:
            cost = input("Enter the cost: ")
        if not revenue:
            revenue = input("Enter the revenue: ")
        if not profit:
            profit = input("Enter the profit: ")
        if not notes:
            notes = input("Enter the notes: ")

if __name__ == "__main__":
    pass
    # Need to add code to call the generate function